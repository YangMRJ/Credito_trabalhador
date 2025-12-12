import pandas as pd
import oracledb
from sqlalchemy import create_engine
from tqdm import tqdm
from colorama import Fore, Style, init
import warnings
import re
import sys
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import shutil
import glob
import calendar

# Configurações iniciais
init(autoreset=True)
warnings.filterwarnings('ignore')

# Configuração do Oracle Client
oracledb.init_oracle_client(lib_dir=r"C:/ORACLE/instantclient_23_7")

# Credenciais do banco de dados
DB_CONFIG = {
    "user": "10987800",
    "password": "senha1",
    "host": "10.70.6.21",
    "port": 1526,
    "service_name": "p01.pcrj"
}

# String de conexão para SQLAlchemy
CONN_STRING = f"oracle+oracledb://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/?service_name={DB_CONFIG['service_name']}"

# Nome do arquivo fixo
ARQUIVO_EXCEL = "ARQUIVO-Consignações-CLT.xlsx"

# Rubricas a serem excluídas do cálculo da margem líquida
RUBRICAS_EXCLUIDAS = [
    1029, 652, 3791, 906, 923, 904, 953, 955, 3515, 3514, 3518, 3777, 3785, 961, 3786,
    3787, 3788, 3523, 3524, 3537, 3538, 3539, 3759, 3760, 3761, 3762, 3790, 905,
    1026, 3781, 3780, 3779, 3778, 3775, 3774, 3769, 3763, 3765, 3766, 3767, 3768
]

def arquivar_relatorio_anterior():
    """
    Verifica se existe um relatório Excel anterior, cria uma pasta baseada no mês/ano
    de modificação do arquivo e o move para lá.
    """
    try:
        archive_base_path = r"C:\Users\03738044\Desktop\Projetos\ATC\CLT\Old"
        report_files = glob.glob('RELATORIO_CONSIGNACOES_*.xlsx')
        
        if not report_files:
            print(f"{Fore.CYAN}Nenhum relatório Excel anterior encontrado para arquivar.{Style.RESET_ALL}")
            return
            
        latest_report = max(report_files, key=os.path.getmtime)
        mod_time = os.path.getmtime(latest_report)
        mod_date = datetime.fromtimestamp(mod_time)
        destination_folder_name = mod_date.strftime("%m.%y")
        destination_folder_path = os.path.join(archive_base_path, destination_folder_name)
        
        os.makedirs(destination_folder_path, exist_ok=True)
        shutil.move(latest_report, destination_folder_path)
        
        print(f"{Fore.GREEN}Relatório Excel anterior '{latest_report}' arquivado em: {destination_folder_path}{Style.RESET_ALL}")

    except Exception as e:
        print(f"{Fore.RED}ERRO ao arquivar relatório Excel anterior: {str(e)}{Style.RESET_ALL}")

def arquivar_csvs_anteriores():
    """
    Encontra todos os CSVs de empresas (comum e zerado), cria uma subpasta para a empresa dentro
    da pasta de arquivamento do mês e move os arquivos para lá.
    """
    try:
        archive_base_path = r"C:\Users\03738044\Desktop\Projetos\ATC\CLT\Old"
        
        # Procura por ambos os padrões de arquivos CSV
        csv_files_comum = glob.glob('RELATORIO_EMPRESA_*.csv')
        csv_files_zerado = glob.glob('RELATORIO_ZERADO_EMPRESA_*.csv')
        all_csv_files = csv_files_comum + csv_files_zerado
        
        if not all_csv_files:
            print(f"{Fore.CYAN}Nenhum arquivo CSV de empresa anterior encontrado para arquivar.{Style.RESET_ALL}")
            return

        total_moved = 0
        for csv_file in all_csv_files:
            try:
                # Extrai o código da empresa do nome do arquivo (funciona para ambos os layouts)
                filename_parts = os.path.basename(csv_file).split('_')
                company_code = filename_parts[2]
                company_folder_name = f"emp_{company_code}"

                # Lógica de arquivamento
                mod_time = os.path.getmtime(csv_file)
                mod_date = datetime.fromtimestamp(mod_time)
                month_year_folder_name = mod_date.strftime("%m.%y")
                
                # Cria o caminho de destino aninhado: Old -> MM.YY -> emp_XX
                destination_folder_path = os.path.join(archive_base_path, month_year_folder_name, company_folder_name)
                
                os.makedirs(destination_folder_path, exist_ok=True)
                shutil.move(csv_file, destination_folder_path)
                total_moved += 1
            except IndexError:
                print(f"{Fore.YELLOW}Aviso: O arquivo '{csv_file}' não segue o padrão de nome esperado e não será arquivado.{Style.RESET_ALL}")
                continue
        
        if total_moved > 0:
            print(f"{Fore.GREEN}Total de {total_moved} CSVs de empresas anteriores foram arquivados.{Style.RESET_ALL}")

    except Exception as e:
        print(f"{Fore.RED}ERRO ao arquivar CSVs anteriores: {str(e)}{Style.RESET_ALL}")


def formatar_matricula(matricula, emp_codigo):
    """Formata a matrícula conforme as regras especificadas."""
    
    # Regras Especiais de Retorno Direto (sem formatação D.DDD.DDD-D)
    if emp_codigo == 1:
        return str(matricula).strip()
    elif emp_codigo == 10:
        return str(matricula).strip()
    
    # --- Início da Lógica de Formatação D.DDD.DDD-D ---
    mat_limpa = re.sub(r'[^0-9]', '', str(matricula))
    
    # Prefixação Baseada no Código da Empresa
    if emp_codigo == 16 and not mat_limpa.startswith('1'):
        mat_limpa = '1' + mat_limpa
    elif (emp_codigo == 18 or emp_codigo == 21) and not mat_limpa.startswith('2'):
        mat_limpa = '2' + mat_limpa
    
    # --- NOVAS REGRAS DE PREFIXAÇÃO ---
    elif emp_codigo == 17 and not mat_limpa.startswith('3'):
        mat_limpa = '3' + mat_limpa
    elif emp_codigo == 14 and not mat_limpa.startswith('2'):
        mat_limpa = '2' + mat_limpa
    elif emp_codigo == 19 and not mat_limpa.startswith('4'):
        mat_limpa = '4' + mat_limpa
    # --- NOVO TRECHO ADICIONADO PARA EMPRESA 15 ---
    elif emp_codigo == 15 and not mat_limpa.startswith('1'):
        mat_limpa = '1' + mat_limpa
    # ----------------------------------
    
    # Formatação Final para 8 dígitos (D.DDD.DDD-D)
    if len(mat_limpa) >= 8:
        return f"{mat_limpa[0]}.{mat_limpa[1:4]}.{mat_limpa[4:7]}-{mat_limpa[7]}"
    else:
        # Completa com zeros à direita se tiver menos de 8 dígitos
        mat_limpa = mat_limpa.ljust(8, '0')
        return f"{mat_limpa[0]}.{mat_limpa[1:4]}.{mat_limpa[4:7]}-{mat_limpa[7]}"
    
def consultar_dados_consignacao(df):
    """Consulta os dados de consignação com margem líquida CORRETA (subtraindo consignações, exceto as rubricas excluídas)."""
    df['MATRICULA_PADRONIZADA'] = df.apply(
        lambda x: formatar_matricula(x['matricula'], x['emp_Codigo']), 
        axis=1
    )
    
    matriculas_consulta = df['MATRICULA_PADRONIZADA'].unique().tolist()
    
    if not matriculas_consulta:
        print(f"{Fore.YELLOW}Aviso: Nenhuma matrícula válida para consulta{Style.RESET_ALL}")
        return {}
    
    resultados = {}
    
    try:
        with oracledb.connect(
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            dsn=f"{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['service_name']}"
        ) as conn:
            with conn.cursor() as cursor:
                cursor.execute("ALTER SESSION SET nls_date_format='DD/MM/YYYY'")

                total_matriculas = len(matriculas_consulta)
                print(f"{Fore.CYAN}Iniciando consulta para {total_matriculas} matrículas...{Style.RESET_ALL}")
                
                with tqdm(total=total_matriculas, desc="Consultando dados", unit="mat") as pbar:
                    for matricula in matriculas_consulta:
                        try:
                            # Verifica status do vínculo
                            query_status = """
                            SELECT 
                                CASE 
                                    WHEN V.DTVAC IS NOT NULL THEN 'DESLIGADO'
                                    WHEN V.DTAPOSENT IS NOT NULL THEN 'APOSENTADO'
                                    ELSE 'ATIVO'
                                END AS STATUS_VINCULO
                            FROM ERGON.VINCULOS V
                            WHERE V.MATRIC = :matricula
                            AND ROWNUM = 1
                            """
                            cursor.execute(query_status, matricula=matricula)
                            status_row = cursor.fetchone()
                            status = status_row[0] if status_row else 'ATIVO'

                            # Se desligado, retorna zeros
                            if status == 'DESLIGADO':
                                resultados[matricula] = {
                                    'STATUS': 'DESLIGADO',
                                    'BASE_1023': 0,
                                    'MARGEM_BRUTA': 0,
                                    'MARGEM_LIQUIDA': 0,
                                    'VALOR_999': 0,
                                    'PLANO_SAUDE': 0
                                }
                                pbar.update(1)
                                continue

                            # Consulta unificada com margem líquida CORRIGIDA
                            query = f"""
                            SELECT 
                                V.MATRIC, 
                                CASE 
                                    WHEN V.DTAPOSENT IS NOT NULL THEN 'APOSENTADO'
                                    ELSE 'ATIVO'
                                END AS STATUS,
                                -- BASE 1023 = (Vantagens - Descontos)
                                (SUM(CASE 
                                    WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                    WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                END)) AS BASE_1023,
                                
                                -- Margem Bruta = BASE 1023 * 35%
                                (SUM(CASE 
                                    WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                    WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                END)) * 0.35 AS MARGEM_BRUTA,
                                
                                -- MARGEM LÍQUIDA = Margem Bruta - OUTRAS CONSIGNAÇÕES (EXCETO RUBRICAS_EXCLUIDAS)
                                (
                                    (SUM(CASE 
                                        WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                        WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                    END)) * 0.35
                                ) - (
                                    SELECT NVL(SUM(C.VALOR), 0)
                                    FROM ERGON.CONS C
                                    JOIN ERGON.RUBRICAS R2 ON R2.RUBRICA = C.RUBRICA
                                    JOIN ERGON.VINCULOS V2 ON V2.NUMFUNC = C.NUMFUNC AND V2.NUMERO = C.NUMVINC
                                    WHERE V2.MATRIC = V.MATRIC
                                    AND R2.TIPORUBR = 'CONSIGNATARIAS'
                                    AND C.DTINI = '01/11/2025'
                                    AND C.RUBRICA NOT IN ({','.join(map(str, RUBRICAS_EXCLUIDAS))})
                                ) AS MARGEM_LIQUIDA,
                                
                                -- Rubrica 999 (opcional)
                                NVL((
                                    SELECT SUM(FFF4.VALOR)
                                    FROM ERGON.FICHAS_FINANCEIRAS FFF4
                                    JOIN ERGON.VINCULOS V4 ON V4.NUMFUNC = FFF4.NUMFUNC AND V4.NUMERO = FFF4.NUMVINC
                                    WHERE V4.MATRIC = V.MATRIC
                                    AND FFF4.RUBRICA = 999
                                    AND FFF4.MES_ANO_FOLHA = '01/11/2025'
                                ), 0) as VALOR_999,
                                
                                -- Valor do plano de saúde (informação adicional) + Rubrica 3564
                                NVL((
                                    SELECT SUM(FFF5.VALOR)
                                    FROM ERGON.FICHAS_FINANCEIRAS FFF5
                                    JOIN ERGON.VINCULOS V5 ON V5.NUMFUNC = FFF5.NUMFUNC AND V5.NUMERO = FFF5.NUMVINC
                                    JOIN ERGON.RUBRICAS R5 ON R5.RUBRICA = FFF5.RUBRICA
                                    WHERE V5.MATRIC = V.MATRIC
                                    AND (R5.FLEX_CAMPO_10 = 'S' OR R5.RUBRICA = 3564) -- <-- ALTERAÇÃO AQUI
                                    AND FFF5.MES_ANO_FOLHA = '01/11/2025'
                                ), 0) as PLANO_SAUDE
                            FROM ERGON.FATORES_RUBRICA_GERAL FTR
                            JOIN ERGON.FICHAS_FINANCEIRAS FFF ON FTR.RUBRICA = FFF.RUBRICA
                            JOIN ERGON.VINCULOS V ON V.NUMFUNC = FFF.NUMFUNC AND V.NUMERO = FFF.NUMVINC
                            JOIN ERGON.RUBRICAS R ON FFF.RUBRICA = R.RUBRICA
                            JOIN ERGON.FOLHAS_EMP fe ON FFF.NUM_FOLHA = fe.NUMERO AND FFF.emp_codigo = fe.EMP_CODIGO AND fe.MES_ANO = FFF.MES_ANO_FOLHA
                            WHERE FTR.FATOR IN ('CRED MARGEM CONS')
                            AND FTR.DTFIM IS NULL
                            AND FFF.MES_ANO_FOLHA = '01/11/2025'
                            AND FFF.EMP_CODIGO IN (1, 10, 13, 14, 15, 16, 17, 18, 19, 21, 23)
                            AND FFF.EMP_CODIGO < 80
                            AND fe.TIPO_FOLHA = 'NORMAL'
                            AND v.matric = :matricula
                            GROUP BY V.MATRIC, V.DTAPOSENT
                            """
                            cursor.execute(query, matricula=matricula)
                            row = cursor.fetchone()
                            
                            pbar.update(1)

                            if row:
                                resultados[matricula] = {
                                    'STATUS': row[1],
                                    'BASE_1023': row[2] if row[2] else 0,
                                    'MARGEM_BRUTA': row[3] if row[3] else 0,
                                    'MARGEM_LIQUIDA': row[4] if row[4] else 0,
                                    'VALOR_999': row[5] if row[5] else 0,
                                    'PLANO_SAUDE': row[6] if row[6] else 0
                                }

                        except oracledb.DatabaseError as e:
                            error, = e.args
                            print(f"\n{Fore.RED}Erro na matrícula {matricula}:")
                            print(f"Código: {error.code}")
                            print(f"Mensagem: {error.message}{Style.RESET_ALL}")
                            continue

    except Exception as e:
        print(f"{Fore.RED}Erro na conexão: {str(e)}{Style.RESET_ALL}")
        return {}

    print(f"\n{Fore.GREEN}Consulta finalizada. Valores encontrados: {len(resultados)}{Style.RESET_ALL}")
    return resultados

def calcular_status_desconto(df):
    """Calcula o status de desconto e o valor descontado para cada contrato, respeitando a ordem cronológica."""
    # Garantir que a data está no formato correto
    df['dataInicioContrato'] = pd.to_datetime(df['dataInicioContrato'], format='%d/%m/%Y', errors='coerce')
    
    # Ordenar por matrícula e data de início do contrato (do mais antigo para o mais recente)
    df = df.sort_values(by=['MATRICULA_PADRONIZADA', 'dataInicioContrato'], ascending=[True, True])
    
    # Inicializar colunas de saída
    df['STATUS_DESCONTO'] = 'SEM DESCONTO'
    df['VALOR_DESCONTADO'] = 0.0
    
    # Processar cada matrícula
    for matricula in df['MATRICULA_PADRONIZADA'].unique():
        # Filtrar registros da matrícula atual
        df_matricula = df[df['MATRICULA_PADRONIZADA'] == matricula].copy()
        
        # Obter margem líquida (assume-se que é a mesma para todos os contratos da matrícula)
        margem_liquida = df_matricula['MARGEM LÍQUIDA 35% (R$)'].iloc[0]
        
        # Processar cada contrato na ordem cronológica
        margem_restante = margem_liquida if margem_liquida > 0 else 0
        for idx in df_matricula.index:
            parcela = df.at[idx, 'PARCELA (R$)']
            
            if margem_restante <= 0:
                # Sem margem disponível, manter como SEM DESCONTO
                df.at[idx, 'STATUS_DESCONTO'] = 'SEM DESCONTO'
                df.at[idx, 'VALOR_DESCONTADO'] = 0.0
            elif parcela <= margem_restante:
                # Parcela cabe na margem restante
                df.at[idx, 'STATUS_DESCONTO'] = 'DESCONTO COMPLETO'
                df.at[idx, 'VALOR_DESCONTADO'] = parcela
                margem_restante -= parcela
            else:
                # Parcela excede a margem restante, desconto parcial
                df.at[idx, 'STATUS_DESCONTO'] = 'DESCONTO PARCIAL'
                df.at[idx, 'VALOR_DESCONTADO'] = margem_restante
                margem_restante = 0
    
    return df

def processar_arquivo():
    try:
        # Arquiva os relatórios (Excel e CSVs) da execução anterior
        arquivar_relatorio_anterior()
        arquivar_csvs_anteriores()

        # Ler arquivo Excel
        print(f"\n{Fore.YELLOW}Processando arquivo: {ARQUIVO_EXCEL}{Style.RESET_ALL}")
        df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Planilha1')
        
        colunas_necessarias = ['contrato', 'cpf', 'matricula', 'emp_Codigo', 'valorParcela', 'nomeTrabalhador', 'dataInicioContrato']
        for coluna in colunas_necessarias:
            if coluna not in df.columns:
                raise ValueError(f"{Fore.RED}Coluna obrigatória '{coluna}' não encontrada{Style.RESET_ALL}")
        
        # Consultar dados no Oracle
        dados_por_matricula = consultar_dados_consignacao(df)
        
        if not dados_por_matricula:
            raise ValueError(f"{Fore.RED}Não foi possível obter os dados do Oracle{Style.RESET_ALL}")
        
        # Adicionar dados ao DataFrame
        df['MATRICULA_PADRONIZADA'] = df.apply(
            lambda x: formatar_matricula(x['matricula'], x['emp_Codigo']), 
            axis=1
        )
        
        df['STATUS'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_por_matricula.get(x, {}).get('STATUS', 'ATIVO'))
        df['BASE_1023'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_por_matricula.get(x, {}).get('BASE_1023', 0))
        df['MARGEM_BRUTA'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_por_matricula.get(x, {}).get('MARGEM_BRUTA', 0))
        df['MARGEM_LIQUIDA_35'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_por_matricula.get(x, {}).get('MARGEM_LIQUIDA', 0))
        df['VALOR_999'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_por_matricula.get(x, {}).get('VALOR_999', 0))
        
        df = df[df['BASE_1023'].notna()]
        
        if df.empty:
            raise ValueError(f"{Fore.RED}Nenhum registro com dados válidos{Style.RESET_ALL}")
        
        # Renomear colunas para padronização
        df = df.rename(columns={
            'valorParcela': 'PARCELA (R$)',
            'MARGEM_LIQUIDA_35': 'MARGEM LÍQUIDA 35% (R$)'
        })
        
        # Define a data de processamento (deve ser a mesma usada na query)
        processing_date = datetime.strptime('01/11/2025', '%d/%m/%Y')
        
        # Calcular status de desconto
        df = calcular_status_desconto(df)

        # --- NOVA LÓGICA DO COMPLEMENTO ---
        # Como o DataFrame já está ordenado por matrícula e data do contrato,
        # podemos simplesmente criar uma contagem cumulativa dentro de cada grupo de matrícula.
        df['COMPLEMENTO'] = 'EMPRESTIMO_' + (df.groupby('MATRICULA_PADRONIZADA').cumcount() + 1).astype(str).str.zfill(2)
        
        # Relatório final em Excel
        colunas_relatorio = [
            'contrato', 'cpf', 'MATRICULA_PADRONIZADA', 'nomeTrabalhador', 'STATUS', 'BASE_1023', 'PARCELA (R$)',
            'MARGEM LÍQUIDA 35% (R$)', 'valorEmprestimo', 'emp_Codigo',
            'STATUS_DESCONTO', 'VALOR_DESCONTADO', 'dataInicioContrato', 'COMPLEMENTO'
        ]
        
        colunas_disponiveis = [col for col in colunas_relatorio if col in df.columns]
        df_relatorio = df[colunas_disponiveis]
        
        mapeamento_nomes = {
            'contrato': 'contrato',
            'cpf': 'cpf',
            'MATRICULA_PADRONIZADA': 'matricula',
            'nomeTrabalhador': 'nome',
            'STATUS': 'status',
            'BASE_1023': 'base 1023',
            'PARCELA (R$)': 'parcela',
            'MARGEM LÍQUIDA 35% (R$)': 'margem liquida',
            'valorEmprestimo': 'emprestimo',
            'emp_Codigo': 'cod_empresa',
            'STATUS_DESCONTO': 'status desconto',
            'VALOR_DESCONTADO': 'valor descontado',
            'dataInicioContrato': 'data inicio',
            'COMPLEMENTO': 'complemento'
        }
        df_relatorio = df_relatorio.rename(columns=mapeamento_nomes)
        
        # Formatação de valores
        colunas_numericas = [
            'base 1023', 'parcela', 'margem liquida', 'emprestimo',
            'valor descontado'
        ]
        
        for coluna in colunas_numericas:
            if coluna in df_relatorio.columns:
                df_relatorio[coluna] = df_relatorio[coluna].apply(lambda x: round(float(x), 2) if pd.notnull(x) else x)
        
        # Ordenar resultados
        df_relatorio = df_relatorio.sort_values(
            ['matricula', 'data inicio'],
            ascending=[True, True]
        )
        
        # Exibir estatísticas
        print(f"\n{Fore.GREEN}RELATÓRIO COMPLETO DE CONSIGNAÇÕES{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Total de registros: {len(df_relatorio)}{Style.RESET_ALL}")
        
        # Exportar para Excel com formatação condicional
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_relatorio = f"RELATORIO_CONSIGNACOES_{timestamp}.xlsx"
        df_relatorio.to_excel(nome_relatorio, sheet_name='REGISTROS', index=False)
        
        # Aplicar formatação condicional
        wb = load_workbook(nome_relatorio)
        ws = wb.active
        
        verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        azul = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        amarelo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        status_col_idx = df_relatorio.columns.get_loc('status') + 1
        status_desconto_col_idx = df_relatorio.columns.get_loc('status desconto') + 1
        
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=status_col_idx).value
            status_desconto = ws.cell(row=row, column=status_desconto_col_idx).value
            
            fill_color = amarelo
            if status in ['DESLIGADO', 'APOSENTADO']:
                fill_color = vermelho
            elif status_desconto == 'DESCONTO COMPLETO':
                fill_color = verde
            elif status_desconto == 'DESCONTO PARCIAL':
                fill_color = azul
            
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill_color
        
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(nome_relatorio)
        
        print(f"\n{Fore.GREEN}Relatório principal salvo como: {nome_relatorio}{Style.RESET_ALL}")

        # --- Gera arquivos CSV (Layout Comum e Layout Zerado) para TODAS as empresas ---
        print(f"\n{Fore.CYAN}Gerando layouts CSV por empresa...{Style.RESET_ALL}")
        
        codigos_empresa = df_relatorio['cod_empresa'].unique()
        
        # Calcula as datas de início e fim do mês de processamento
        data_inicio_str = processing_date.strftime('01/%m/%Y')
        _, num_dias = calendar.monthrange(processing_date.year, processing_date.month)
        data_fim_str = f"{num_dias}/{processing_date.month:02d}/{processing_date.year}"

        for codigo in codigos_empresa:
            df_empresa = df_relatorio[df_relatorio['cod_empresa'] == codigo]
            
            # --- NOVO FILTRO: Usa apenas registros com desconto completo ou parcial ---
            df_filtrado = df_empresa[df_empresa['status desconto'].isin(['DESCONTO COMPLETO', 'DESCONTO PARCIAL'])].copy()

            if df_filtrado.empty:
                print(f"{Fore.MAGENTA}Nenhum registro com desconto encontrado para a empresa {codigo}. Nenhum CSV será gerado.{Style.RESET_ALL}")
                continue

            # --- 1. GERAÇÃO DO LAYOUT COMUM ---
            df_layout_comum = pd.DataFrame()
            df_layout_comum['matricula'] = df_filtrado['matricula']
            df_layout_comum['data inicio'] = data_inicio_str
            df_layout_comum['data fim'] = data_fim_str
            df_layout_comum['rubrica'] = 1029
            df_layout_comum['complemento'] = df_filtrado['complemento'].fillna('')
            df_layout_comum['tipo'] = 'Livre'
            df_layout_comum['valor descontado'] = df_filtrado['valor descontado']
            df_layout_comum['empresa'] = df_filtrado['cod_empresa']
            
            nome_csv_comum = f"RELATORIO_EMPRESA_{codigo}_{timestamp}.csv"
            # ***** INÍCIO DA MODIFICAÇÃO *****
            df_layout_comum.to_csv(nome_csv_comum, index=False, sep=';', encoding='utf-8-sig', header=False, decimal=',')
            # ***** FIM DA MODIFICAÇÃO *****
            print(f"{Fore.GREEN}Layout Comum para empresa {codigo} salvo como: {nome_csv_comum}{Style.RESET_ALL}")

            # --- 2. GERAÇÃO DO LAYOUT ZERADO ---
            df_layout_zerado = pd.DataFrame()
            df_layout_zerado['matricula'] = df_filtrado['matricula']
            df_layout_zerado['data inicio'] = data_inicio_str
            df_layout_zerado['rubrica'] = 1029
            df_layout_zerado['complemento'] = df_filtrado['complemento'].fillna('')
            df_layout_zerado['tipo'] = 'A'
            df_layout_zerado['valor'] = 0  # Valor sempre zero
            df_layout_zerado['empresa'] = df_filtrado['cod_empresa']

            nome_csv_zerado = f"RELATORIO_ZERADO_EMPRESA_{codigo}_{timestamp}.csv"
            # ***** INÍCIO DA MODIFICAÇÃO *****
            df_layout_zerado.to_csv(nome_csv_zerado, index=False, sep=';', encoding='utf-8-sig', header=False, decimal=',')
            # ***** FIM DA MODIFICAÇÃO *****
            print(f"{Fore.YELLOW}Layout Zerado para empresa {codigo} salvo como: {nome_csv_zerado}{Style.RESET_ALL}")


    except Exception as e:
        print(f"\n{Fore.RED}ERRO: {str(e)}{Style.RESET_ALL}")
        return -1

if __name__ == "__main__":
    processar_arquivo()