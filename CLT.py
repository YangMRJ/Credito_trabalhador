import pandas as pd
import oracledb
from sqlalchemy import create_engine
from tqdm import tqdm
from colorama import Fore, Style, init
import warnings
import re
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import shutil
import glob
import calendar

# --- AJUSTE PARA MODO THIN E PYINSTALLER ---
# Importar explicitamente para garantir que o pacote de criptografia vá pro .exe
try:
    import cryptography
    from cryptography.hazmat.primitives import hashes
except ImportError:
    pass

# --- CONFIGURAÇÕES GERAIS ---
init(autoreset=True)
warnings.filterwarnings('ignore')

# LISTA DE EMPRESAS (Pastas que serão verificadas/criadas)
LISTA_EMPRESAS = [1, 10, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23]

# --- DATA DINÂMICA (CORRIGIDO PARA MÊS ATUAL) ---
def obter_competencia():
    """
    Calcula automaticamente a competência (1º dia do mês ATUAL).
    Se hoje é 10/02/2026 -> Retorna '01/02/2026'
    """
    hoje = datetime.now()
    mes = hoje.month # Pega o mês atual
    ano = hoje.year
    
    data_formatada = f"01/{mes:02d}/{ano}"
    return data_formatada

DATA_PROCESSAMENTO = obter_competencia()

# --- CONFIGURAÇÃO INTELIGENTE DO ORACLE CLIENT ---
def configurar_oracle():
    """
    Configura o uso EXCLUSIVO do Oracle Client Thick a partir da pasta local.
    Se a pasta não existir ou o carregamento falhar, o script é interrompido.
    """
    try:
        # Detecta o diretório base (seja .py ou .exe)
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))

        # Caminho absoluto para a pasta 'oracle_client'
        caminho_cliente = os.path.abspath(os.path.join(base_dir, "oracle_client"))

        # Verifica obrigatoriedade da pasta
        if not os.path.exists(caminho_cliente):
            print(f"{Fore.RED}ERRO CRÍTICO: Pasta 'oracle_client' não encontrada em: {caminho_cliente}{Style.RESET_ALL}")
            print(f"{Fore.YELLOW}Certifique-se de que a pasta está no mesmo local que o script/executável.{Style.RESET_ALL}")
            input("Pressione ENTER para sair...")
            sys.exit(1)

        # Adiciona o caminho ao PATH do sistema operacional (importante para DLLs no Windows)
        os.environ["PATH"] = caminho_cliente + os.pathsep + os.environ.get("PATH", "")

        # Inicializa o cliente Thick apontando explicitamente para a pasta
        oracledb.init_oracle_client(lib_dir=caminho_cliente)
        
        print(f"{Fore.GREEN}Oracle Client (Modo Thick) carregado com sucesso de:{Style.RESET_ALL}")
        print(f"{Fore.CYAN}  -> {caminho_cliente}{Style.RESET_ALL}")

    except Exception as e:
        print(f"{Fore.RED}ERRO FATAL AO INICIALIZAR ORACLE CLIENT: {e}{Style.RESET_ALL}")
        input("Pressione ENTER para sair...")
        sys.exit(1)

# Chama a configuração logo no início
configurar_oracle()

# Credenciais do banco de dados
DB_CONFIG = {
    "user": "10987800",
    "password": "senha1",
    "host": "10.70.6.21",
    "port": 1526,
    "service_name": "p01.pcrj"
}

RUBRICAS_EXCLUIDAS = [
    1029, 652, 3791, 906, 923, 904, 953, 955, 3515, 3514, 3518, 3777, 3785, 961, 3786,
    3787, 3788, 3523, 3524, 3537, 3538, 3539, 3759, 3760, 3761, 3762, 3790, 905,
    1026, 3781, 3780, 3779, 3778, 3775, 3774, 3769, 3763, 3765, 3766, 3767, 3768
]

def formatar_matricula(matricula, emp_codigo):
    """Formata a matrícula conforme as regras da empresa e trata excesso de zeros."""
    try:
        emp_codigo = int(emp_codigo)
    except:
        return str(matricula)
    
    mat_str = str(matricula).strip()
    
    # Empresas 1 e 10 não seguem o padrão de máscara de pontos
    if (emp_codigo == 1 or emp_codigo == 10):
        return mat_str
    
    # Remove qualquer caractere não numérico
    mat_limpa = re.sub(r'[^0-9]', '', mat_str)
    
    # Trata matrículas com zeros à esquerda (ex: de sistemas externos)
    # Se após limpar tiver mais de 8 dígitos, remove os zeros iniciais
    if len(mat_limpa) > 8:
        mat_limpa = mat_limpa.lstrip('0')
    
    # Se a matrícula estiver vazia após o strip, retorna original para evitar erro
    if not mat_limpa: return mat_str

    # Aplica prefixos específicos por grupo de empresa
    if (emp_codigo in [13, 15, 16, 20]) and not mat_limpa.startswith('1'):
        mat_limpa = '1' + mat_limpa
    elif (emp_codigo in [14, 18, 21]) and not mat_limpa.startswith('2'):
        mat_limpa = '2' + mat_limpa
    elif emp_codigo == 17 and not mat_limpa.startswith('3'):
        mat_limpa = '3' + mat_limpa
    elif (emp_codigo in [19, 23]) and not mat_limpa.startswith('4'):
        mat_limpa = '4' + mat_limpa

    # Aplica a máscara X.XXX.XXX-X
    if len(mat_limpa) >= 8:
        return f"{mat_limpa[0]}.{mat_limpa[1:4]}.{mat_limpa[4:7]}-{mat_limpa[7]}"
    else:
        # Preenche com zeros à direita se for menor que 8 (raro)
        mat_limpa = mat_limpa.ljust(8, '0')
        return f"{mat_limpa[0]}.{mat_limpa[1:4]}.{mat_limpa[4:7]}-{mat_limpa[7]}"

def corrigir_matriculas_por_cpf(df, emp_codigo):
    """
    Verifica se a matrícula do arquivo pertence ao CPF. 
    Se o par for válido, mantém (mesmo desligado). 
    Caso contrário, busca a matrícula ativa do CPF.
    """
    print(f"  > Verificando/Corrigindo matrículas no banco de dados...")
    
    if 'cpf' in df.columns:
        df['cpf_limpo'] = df['cpf'].astype(str).str.replace(r'[^0-9]', '', regex=True)
        
        try:
            with oracledb.connect(
                user=DB_CONFIG['user'],
                password=DB_CONFIG['password'],
                dsn=f"{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['service_name']}"
            ) as conn:
                with conn.cursor() as cursor:
                    
                    for idx, row in df.iterrows():
                        matricula_original = str(row['matricula']).strip()
                        cpf_atual = row['cpf_limpo']
                        
                        # Formata a matrícula para o padrão do banco (X.XXX.XXX-X)
                        matricula_formatada = formatar_matricula(matricula_original, emp_codigo)
                        
                        # QUERY 1: Valida se o CPF é dono dessa matrícula na empresa (Ativo ou não)
                        # Usando apelidos longos para evitar conflitos de nomes no Oracle
                        query_valida = """
                            SELECT COUNT(*) 
                            FROM ERGON.VINCULOS VINC_VAL
                            JOIN ERGON.FUNCIONARIOS FUNC_VAL ON VINC_VAL.NUMFUNC = FUNC_VAL.NUMERO
                            WHERE (VINC_VAL.MATRIC = :mat_form OR VINC_VAL.MATRIC = :mat_orig)
                            AND VINC_VAL.EMP_CODIGO = :emp_codigo
                            AND FUNC_VAL.CPF = :cpf
                        """
                        cursor.execute(query_valida, 
                                       mat_form=matricula_formatada, 
                                       mat_orig=matricula_original,
                                       emp_codigo=emp_codigo, 
                                       cpf=cpf_atual)
                        
                        if cursor.fetchone()[0] > 0:
                            # Se o par CPF+Matrícula existe, mantemos a escolha do arquivo
                            continue
                        
                        # QUERY 2: Se a matrícula não pertence ao CPF, busca a matrícula 
                        # mais recente (DTVAC NULLS FIRST prioriza as ativas)
                        # Corrigido: Removido DTADMISS que causava erro de identificador
                        query_busca = """
                            SELECT MATRIC FROM (
                                SELECT VINC_BUSCA.MATRIC 
                                FROM ERGON.FUNCIONARIOS FUNC_BUSCA
                                JOIN ERGON.VINCULOS VINC_BUSCA ON VINC_BUSCA.NUMFUNC = FUNC_BUSCA.NUMERO
                                WHERE FUNC_BUSCA.CPF = :cpf
                                AND VINC_BUSCA.EMP_CODIGO = :emp_codigo
                                ORDER BY VINC_BUSCA.DTVAC NULLS FIRST
                            ) WHERE ROWNUM = 1
                        """
                        cursor.execute(query_busca, cpf=cpf_atual, emp_codigo=emp_codigo)
                        res = cursor.fetchone()
                        
                        if res:
                            mat_correta = res[0]
                            df.at[idx, 'matricula'] = mat_correta
                            print(f"{Fore.YELLOW}      - CPF {cpf_atual}: Matrícula '{matricula_original}' corrigida para '{mat_correta}'{Style.RESET_ALL}")
                        else:
                            print(f"{Fore.RED}      - CPF {cpf_atual}: Matrícula '{matricula_original}' não localizada na Empresa {emp_codigo}.{Style.RESET_ALL}")
                                
        except Exception as e:
             print(f"{Fore.RED}    Erro Crítico na Correção por CPF: {e}{Style.RESET_ALL}")
             
    return df

def consultar_dados_consignacao(df, data_ref):
    """Consulta os dados de consignação no Oracle."""
    df['MATRICULA_PADRONIZADA'] = df.apply(
        lambda x: formatar_matricula(x['matricula'], x['emp_Codigo']), 
        axis=1
    )
    
    matriculas_consulta = df['MATRICULA_PADRONIZADA'].unique().tolist()
    
    if not matriculas_consulta:
        return {}
    
    resultados = {}
    
    try:
        with oracledb.connect(
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            dsn=f"{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['service_name']}"
        ) as conn:
            with conn.cursor() as cursor:
                cursor.execute("ALTER SESSION SET nls_date_format='DD/MM/YYYY'")

                total_matriculas = len(matriculas_consulta)
                print(f"   -> Consultando {total_matriculas} matrículas no Oracle...")
                
                query = f"""
                            SELECT 
                                V.MATRIC,
                                V.NUMFUNC,
                                V.NUMERO AS NUMVINC,
                                CASE 
                                    WHEN V.DTAPOSENT IS NOT NULL THEN 'APOSENTADO'
                                    ELSE 'ATIVO'
                                END AS STATUS,
                                -- BASE 1023
                                (SUM(CASE 
                                    WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                    WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                END)) AS BASE_1023,
                                
                                -- Margem Bruta
                                (SUM(CASE 
                                    WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                    WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                END)) * 0.35 AS MARGEM_BRUTA,
                                
                                -- MARGEM LÍQUIDA
                                (
                                    (SUM(CASE 
                                        WHEN FTR.SINAL = -1 THEN FFF.VALOR * (-1) 
                                        WHEN FTR.SINAL <> -1 AND R.TIPORUBR = 'VANTAGENS' THEN FFF.VALOR 
                                    END)) * 0.35
                                ) - (
                                    SELECT NVL(SUM(C.VALOR), 0)
                                    FROM ERGON.CONS C
                                    JOIN ERGON.RUBRICAS R2 ON R2.RUBRICA = C.RUBRICA
                                    JOIN ERGON.VINCULOS V2 ON V2.NUMFUNC = C.NUMFUNC AND V2.NUMERO = C.NUMVINC
                                    WHERE V2.MATRIC = V.MATRIC
                                    AND R2.TIPORUBR = 'CONSIGNATARIAS'
                                    AND C.DTINI = :data_ref
                                    AND C.RUBRICA NOT IN ({','.join(map(str, RUBRICAS_EXCLUIDAS))})
                                ) AS MARGEM_LIQUIDA,
                                
                                -- Rubrica 999
                                NVL((
                                    SELECT SUM(FFF4.VALOR)
                                    FROM ERGON.FICHAS_FINANCEIRAS FFF4
                                    JOIN ERGON.VINCULOS V4 ON V4.NUMFUNC = FFF4.NUMFUNC AND V4.NUMERO = FFF4.NUMVINC
                                    WHERE V4.MATRIC = V.MATRIC
                                    AND FFF4.RUBRICA = 999
                                    AND FFF4.MES_ANO_FOLHA = :data_ref
                                ), 0) as VALOR_999,
                                
                                -- Plano de saúde
                                NVL((
                                    SELECT SUM(FFF5.VALOR)
                                    FROM ERGON.FICHAS_FINANCEIRAS FFF5
                                    JOIN ERGON.VINCULOS V5 ON V5.NUMFUNC = FFF5.NUMFUNC AND V5.NUMERO = FFF5.NUMVINC
                                    JOIN ERGON.RUBRICAS R5 ON R5.RUBRICA = FFF5.RUBRICA
                                    WHERE V5.MATRIC = V.MATRIC
                                    AND (R5.FLEX_CAMPO_10 = 'S' OR R5.RUBRICA = 3564)
                                    AND FFF5.MES_ANO_FOLHA = :data_ref
                                ), 0) as PLANO_SAUDE
                            FROM ERGON.FATORES_RUBRICA_GERAL FTR
                            JOIN ERGON.FICHAS_FINANCEIRAS FFF ON FTR.RUBRICA = FFF.RUBRICA
                            JOIN ERGON.VINCULOS V ON V.NUMFUNC = FFF.NUMFUNC AND V.NUMERO = FFF.NUMVINC
                            JOIN ERGON.RUBRICAS R ON FFF.RUBRICA = R.RUBRICA
                            JOIN ERGON.FOLHAS_EMP fe ON FFF.NUM_FOLHA = fe.NUMERO AND FFF.emp_codigo = fe.EMP_CODIGO AND fe.MES_ANO = FFF.MES_ANO_FOLHA
                            WHERE FTR.FATOR IN ('CRED MARGEM CONS')
                            AND FTR.DTFIM IS NULL
                            AND FFF.MES_ANO_FOLHA = :data_ref
                            AND FFF.EMP_CODIGO IN (1, 10, 13, 14, 15, 16, 17, 18, 19, 20, 21, 23)
                            AND FFF.EMP_CODIGO < 80
                            AND fe.TIPO_FOLHA = 'NORMAL'
                            AND v.matric = :matricula
                            GROUP BY V.MATRIC, V.DTAPOSENT, V.NUMFUNC, V.NUMERO
                            """

                for matricula in matriculas_consulta:
                    try:
                        cursor.execute("SELECT CASE WHEN DTVAC IS NOT NULL THEN 'DESLIGADO' ELSE 'ATIVO' END FROM ERGON.VINCULOS WHERE MATRIC = :m AND ROWNUM=1", m=matricula)
                        status_chk = cursor.fetchone()
                        
                        if status_chk and status_chk[0] == 'DESLIGADO':
                             resultados[matricula] = {'STATUS': 'DESLIGADO', 'BASE_1023': 0, 'MARGEM_LIQUIDA': 0}
                             continue

                        cursor.execute(query, matricula=matricula, data_ref=data_ref)
                        row = cursor.fetchone()
                        
                        if row:
                            resultados[matricula] = {
                                'NUMFUNC': row[1], 'NUMVINC': row[2], 'STATUS': row[3],
                                'BASE_1023': row[4] or 0, 'MARGEM_BRUTA': row[5] or 0,
                                'MARGEM_LIQUIDA': row[6] or 0, 'VALOR_999': row[7] or 0,
                                'PLANO_SAUDE': row[8] or 0
                            }
                        else:
                             resultados[matricula] = {'STATUS': 'NAO_ENCONTRADO_FOLHA'} 

                    except Exception as e:
                        continue

    except Exception as e:
        print(f"{Fore.RED}Erro Conexão Oracle: {e}{Style.RESET_ALL}")
        return {}
    
    return resultados

def calcular_status_desconto(df):
    """Lógica de cálculo de margem e descontos."""
    df['dataInicioContrato'] = pd.to_datetime(df['dataInicioContrato'], format='%d/%m/%Y', errors='coerce')
    df = df.sort_values(by=['MATRICULA_PADRONIZADA', 'dataInicioContrato'], ascending=[True, True])
    
    df['STATUS_DESCONTO'] = 'SEM DESCONTO'
    df['VALOR_DESCONTADO'] = 0.0
    
    for matricula in df['MATRICULA_PADRONIZADA'].unique():
        df_matricula = df[df['MATRICULA_PADRONIZADA'] == matricula].copy()
        try:
            margem_liquida = df_matricula['MARGEM LÍQUIDA 35% (R$)'].iloc[0]
            if pd.isna(margem_liquida): margem_liquida = 0
        except:
            margem_liquida = 0
        
        margem_restante = margem_liquida if margem_liquida > 0 else 0
        
        for idx in df_matricula.index:
            parcela = df.at[idx, 'PARCELA (R$)']
            if pd.isna(parcela): parcela = 0
            
            if margem_restante <= 0:
                df.at[idx, 'STATUS_DESCONTO'] = 'SEM DESCONTO'
                df.at[idx, 'VALOR_DESCONTADO'] = 0.0
            elif parcela <= margem_restante:
                df.at[idx, 'STATUS_DESCONTO'] = 'DESCONTO COMPLETO'
                df.at[idx, 'VALOR_DESCONTADO'] = parcela
                margem_restante -= parcela
            else:
                df.at[idx, 'STATUS_DESCONTO'] = 'DESCONTO PARCIAL'
                df.at[idx, 'VALOR_DESCONTADO'] = margem_restante
                margem_restante = 0
    return df

def arquivar_outputs_antigos(caminho_pasta, emp_codigo, pasta_antigos):
    """
    Procura por Relatórios Finais, Arquivos de Carga, Logs e Rejeitos 
    antigos na pasta e move para 'antigos'.
    """
    print(f"  > Verificando arquivos de saída antigos (Relatórios/Cargas/Logs/Rejeitos)...")
    
    # Adicionamos .rej e .log à lista de busca
    padroes = [
        f"RELATORIO_FINAL_EMP_{emp_codigo}_*.xlsx",
        f"CARGA_ERGON_EMP_{emp_codigo}_*.txt",
        f"CARGA_ERGON_EMP_{emp_codigo}_*.rej",
        f"CARGA_ERGON_EMP_{emp_codigo}_*.log"
    ]
    
    movidos = 0
    for padrao in padroes:
        arquivos_velhos = glob.glob(os.path.join(caminho_pasta, padrao))
        for arquivo in arquivos_velhos:
            try:
                nome_arquivo = os.path.basename(arquivo)
                destino = os.path.join(pasta_antigos, nome_arquivo)
                
                if os.path.exists(destino):
                    os.remove(destino)
                    
                shutil.move(arquivo, destino)
                movidos += 1
            except Exception as e:
                print(f"{Fore.RED}    Erro ao arquivar antigo {os.path.basename(arquivo)}: {e}{Style.RESET_ALL}")
    
    if movidos > 0:
        print(f"    -> {movidos} arquivos (incluindo .rej/.log) movidos para pasta 'antigos'.")

def processar_uma_empresa(emp_codigo, caminho_pasta):
    """Processa todos os arquivos Excel e CSV encontrados na pasta da empresa."""
    
    pasta_antigos = os.path.join(caminho_pasta, "antigos")
    os.makedirs(pasta_antigos, exist_ok=True)

    # --- PASSO 1: LIMPEZA DE ARQUIVOS GERADOS ANTIGOS ---
    arquivar_outputs_antigos(caminho_pasta, emp_codigo, pasta_antigos)

    # --- PASSO 2: ENCONTRAR NOVOS ARQUIVOS DE ENTRADA (XLSX e CSV) ---
    # Procura XLSX e CSV
    todos_arquivos = glob.glob(os.path.join(caminho_pasta, "*.xlsx")) + \
                     glob.glob(os.path.join(caminho_pasta, "*.csv"))
    
    arquivos_validos = []
    
    # FILTRO DE SEGURANÇA: Remove arquivos que parecem ser Outputs (Relatórios ou Cargas)
    for f in todos_arquivos:
        nome = os.path.basename(f).upper()
        # Se começar com RELATORIO ou CARGA, a gente ignora e avisa
        if nome.startswith("RELATORIO_FINAL") or nome.startswith("CARGA_ERGON"):
            # Opcional: print(f"    -> Ignorando arquivo de sistema: {nome}")
            continue
        arquivos_validos.append(f)

    if not arquivos_validos:
        print(f"{Fore.LIGHTBLACK_EX}Empresa {emp_codigo}: Nenhum arquivo de entrada (XLSX/CSV) novo.{Style.RESET_ALL}")
        return

    print(f"{Fore.YELLOW}Empresa {emp_codigo}: Encontrados {len(arquivos_validos)} arquivos de entrada.{Style.RESET_ALL}")

    for arquivo in arquivos_validos:
        nome_arq = os.path.basename(arquivo)
        extensao = os.path.splitext(nome_arq)[1].lower()
        print(f"  > Processando entrada: {nome_arq}")
        
        try:
            # --- LEITURA DO ARQUIVO (MODIFICADO PARA SUPORTAR CSV) ---
            if extensao == '.xlsx':
                df = pd.read_excel(arquivo)
            elif extensao == '.csv':
                # Tenta ler CSV padrão Brasil (Ponto e vírgula e decimal com vírgula)
                try:
                    df = pd.read_csv(arquivo, sep=';', encoding='latin-1', decimal=',')
                except:
                    # Se falhar, tenta formato internacional (vírgula e utf-8)
                    print(f"{Fore.CYAN}    Aviso: Tentando ler CSV com separador vírgula e UTF-8...{Style.RESET_ALL}")
                    df = pd.read_csv(arquivo, sep=',', encoding='utf-8', decimal='.')

            df['emp_Codigo'] = int(emp_codigo)
            
            # Remove espaços em branco dos nomes das colunas (sanity check)
            df.columns = df.columns.str.strip()

            cols_req = ['contrato', 'cpf', 'matricula', 'valorParcela', 'nomeTrabalhador']
            
            # Verifica colunas
            if not all(col in df.columns for col in cols_req):
                colunas_faltantes = [col for col in cols_req if col not in df.columns]
                print(f"{Fore.RED}    Erro: Colunas faltando em {nome_arq}: {colunas_faltantes}{Style.RESET_ALL}")
                continue

            # --- CORREÇÃO HÍBRIDA (PONTO E VÍRGULA) ---
            def converter_valor_hibrido(val):
                """
                Converte o valor de forma inteligente:
                - Se tiver vírgula, assume formato BR (1.000,50) -> Remove ponto milhar, troca vírgula por ponto.
                - Se NÃO tiver vírgula, assume que o ponto já é decimal (1000.50).
                """
                try:
                    # Se já for número (float/int), retorna direto
                    if isinstance(val, (int, float)):
                        return float(val)
                    
                    val_str = str(val).strip()
                    
                    # Lógica: Se tem vírgula, assumimos formato Brasileiro
                    if ',' in val_str:
                        # Remove o ponto de milhar (1.000,00 -> 1000,00)
                        val_str = val_str.replace('.', '')
                        # Troca vírgula por ponto (1000,00 -> 1000.00)
                        val_str = val_str.replace(',', '.')
                    
                    # Se não tem vírgula, mantemos o ponto original para o float() ler corretamente
                    return float(val_str)
                except:
                    return 0.0

            # Aplica a função de conversão
            df['valorParcela'] = df['valorParcela'].apply(converter_valor_hibrido)
            # ------------------------------------------

            # --- NOVA ETAPA: CORREÇÃO POR CPF ---
            df = corrigir_matriculas_por_cpf(df, emp_codigo)
            # ------------------------------------

            # Consultar Oracle
            dados_oracle = consultar_dados_consignacao(df, DATA_PROCESSAMENTO)
            
            if not dados_oracle:
                print(f"{Fore.RED}    Erro: Falha na consulta Oracle ou sem matrículas válidas.{Style.RESET_ALL}")
                continue

            # Mapeamento de dados
            df['MATRICULA_PADRONIZADA'] = df.apply(lambda x: formatar_matricula(x['matricula'], emp_codigo), axis=1)
            df['STATUS'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_oracle.get(x, {}).get('STATUS', 'NAO_ENC'))
            df['NUMFUNC'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_oracle.get(x, {}).get('NUMFUNC', 0))
            df['NUMVINC'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_oracle.get(x, {}).get('NUMVINC', 0))
            df['BASE_1023'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_oracle.get(x, {}).get('BASE_1023', 0))
            df['MARGEM LÍQUIDA 35% (R$)'] = df['MATRICULA_PADRONIZADA'].map(lambda x: dados_oracle.get(x, {}).get('MARGEM_LIQUIDA', 0))
            
            df = df.rename(columns={'valorParcela': 'PARCELA (R$)'})
            
            # Cálculo
            df = calcular_status_desconto(df)
            df['COMPLEMENTO'] = 'EMPRESTIMO_' + (df.groupby('MATRICULA_PADRONIZADA').cumcount() + 1).astype(str).str.zfill(2)

            # --- GERAÇÃO DOS RELATÓRIOS ---
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 1. Relatório Excel
            nome_relatorio = os.path.join(caminho_pasta, f"RELATORIO_FINAL_EMP_{emp_codigo}_{timestamp}.xlsx")
            
            # --- DEFINIÇÃO DA ORDEM DAS COLUNAS ---
            # Define as colunas fixas principais
            colunas_fixas = [
                'contrato', 'cpf', 'MATRICULA_PADRONIZADA', 'nomeTrabalhador', 'STATUS',
                'BASE_1023', 'PARCELA (R$)', 'MARGEM LÍQUIDA 35% (R$)',
                'STATUS_DESCONTO', 'VALOR_DESCONTADO', 'COMPLEMENTO'
            ]
            
            # Lista inicial com as colunas prioritárias (se existirem)
            colunas_finais = []
            cols_prioridade = ['ifConcessora.codigo', 'ifConcessora.descricao']
            
            # Adiciona as prioritárias primeiro se elas existirem no DataFrame
            for col in cols_prioridade:
                if col in df.columns:
                    colunas_finais.append(col)
            
            # Adiciona as demais fixas
            colunas_finais.extend(colunas_fixas)
            
            # Adiciona outras colunas extras (como numeroInscricaoEmpregador) se existirem
            extras = ['numeroInscricaoEmpregador']
            for e in extras:
                if e in df.columns: colunas_finais.append(e)

            # Filtra apenas as colunas que realmente existem no DF para evitar erro de KeyError
            colunas_existentes = [c for c in colunas_finais if c in df.columns]

            df_export = df[colunas_existentes].copy()
            df_export.to_excel(nome_relatorio, index=False)
            
            # --- FORMATAÇÃO DAS CORES ---
            wb = load_workbook(nome_relatorio)
            ws = wb.active
            
            verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')    # Desconto Completo
            azul = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')     # Desconto Parcial
            amarelo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Sem Desconto
            vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Desligado/Aposentado
            
            idx_status = None
            idx_desc = None
            
            for cell in ws[1]:
                if cell.value == 'STATUS': idx_status = cell.column
                if cell.value == 'STATUS_DESCONTO': idx_desc = cell.column
            
            if idx_status and idx_desc:
                for row in range(2, ws.max_row + 1):
                    val_st = ws.cell(row, idx_status).value
                    val_dc = ws.cell(row, idx_desc).value
                    
                    fill_color = None
                    if val_st in ['DESLIGADO', 'APOSENTADO']:
                        fill_color = vermelho
                    elif val_dc == 'DESCONTO COMPLETO':
                        fill_color = verde
                    elif val_dc == 'DESCONTO PARCIAL':
                        fill_color = azul
                    elif val_dc == 'SEM DESCONTO':
                        fill_color = amarelo
                    
                    if fill_color:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row, col).fill = fill_color
            
            # Ajuste de largura
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = adjusted_width

            wb.save(nome_relatorio)

            # 2. Arquivo de Carga (TXT)
            df_carga = df[df['VALOR_DESCONTADO'] > 0].copy()
            if not df_carga.empty:
                linhas = []
                for _, r in df_carga.iterrows():
                    val_str = f"{r['VALOR_DESCONTADO']:.2f}".replace('.', ',')
                    linha = f"{int(r['NUMFUNC'])};{int(r['NUMVINC'])};{DATA_PROCESSAMENTO};1029;1;{r['COMPLEMENTO']};S;{val_str};U3738044;CARGA;{emp_codigo};"
                    linhas.append(linha)
                
                header = [
                    f'@TABELA=[ERGON][MOVIMENTOS][6.6.4][{datetime.now().strftime("%d/%m/%Y")}]',
                    '@CHAVE=[NUMFUNC][NUMBER][NUMVINC][NUMBER][RUBRICA][NUMBER]',
                    '@TAG INICIO=', '@TAG FIM=', '@SEPARADOR=;', '@FORMATO DATA=DD/MM/YYYY',
                    '@COLUNAS=[NUMFUNC][NUMBER][NUMVINC][NUMBER][MES_ANO_DIREITO][DATE][RUBRICA][NUMBER][CHAVE][NUMBER][COMPLEMENTO][VARCHAR2][TIPO_MOVIMENTO][VARCHAR2][VALOR][NUMBER][RESPONSAVEL][VARCHAR2][OBS][VARCHAR2][EMP_CODIGO][NUMBER]'
                ]
                
                txt_path = os.path.join(caminho_pasta, f"CARGA_ERGON_EMP_{emp_codigo}_{timestamp}.txt")
                with open(txt_path, 'w', encoding='latin-1') as ftxt:
                    ftxt.write("\n".join(header + linhas))

            # 3. Arquivar o original
            destino_original = os.path.join(pasta_antigos, nome_arq)
            if os.path.exists(destino_original):
                os.remove(destino_original)
            
            shutil.move(arquivo, destino_original)
            print(f"{Fore.GREEN}    -> Sucesso! Arquivos gerados e movidos.{Style.RESET_ALL}")

        except Exception as e:
            print(f"{Fore.RED}    -> FALHA CRÍTICA no arquivo {nome_arq}: {e}{Style.RESET_ALL}")
            # import traceback
            # traceback.print_exc()

def main():
    root_dir = os.getcwd()
    
    # Se estiver rodando como .exe, mostra onde ele acha que está
    if getattr(sys, 'frozen', False):
         root_dir = os.path.dirname(sys.executable)
         os.chdir(root_dir) # Garante que está na pasta do exe

    print(f"{Fore.CYAN}=== INICIANDO PROCESSAMENTO AUTOMÁTICO ==={Style.RESET_ALL}")
    print(f"Diretório Raiz: {root_dir}")
    print(f"Data de Competência Calculada: {DATA_PROCESSAMENTO}")
    print("-------------------------------------------------")

    for codigo in LISTA_EMPRESAS:
        dir_empresa = os.path.join(root_dir, str(codigo))
        if not os.path.exists(dir_empresa):
            os.makedirs(dir_empresa)
            print(f"Pasta criada: {codigo} (Vazia)")
            continue
        processar_uma_empresa(codigo, dir_empresa)
    
    print("\n-------------------------------------------------")
    print(f"{Fore.CYAN}=== FIM DO PROCESSAMENTO ==={Style.RESET_ALL}")
    input("Pressione ENTER para fechar...") # Para a janela não sumir no fim

if __name__ == "__main__":
    main()